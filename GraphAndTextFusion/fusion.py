import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import fnmatch
from ai import aiFusion

data_path = 'dataSet/JayMe.xlsx'
save_path = 'afterFusion.xlsx'
picture_path = 'dataSet/JayMe'


def ai(id:str,text:str):
    images = []
    prefix = 'image_' #id，用作图片开头搜寻是否有对应图片
    for file_name in os.listdir(picture_path): # 遍历文件夹中的所有文件
        if (file_name.startswith(id)):
            images.append(picture_path + '/' + file_name)
        # if fnmatch.fnmatch(file_name, prefix + '*'):  # 判断文件名是否以特定字符串开头
            
    if (len(images) != 0 ):        
        return [len(images),aiFusion(images,text)]
    else:
        return [len(images),text]

def writeToExcel(pos:str,content:str):
    return

def openSheet(path:str):
    wb = openpyxl.load_workbook(path) # 打开工作簿
    sheets_names = wb.sheetnames # 取得工作表
    inSheet = wb.active # 获取活动表对应的表对象(表对象就是Worksheet类的对象)
    return inSheet

def main():
    inSheet = openSheet(data_path)
    rowMax = inSheet.max_row #最大行数
    colMax = inSheet.max_column #最大列数
    wb = openpyxl.load_workbook(save_path) # 存储的工作簿
    outSheet = wb.active
    outSheet['A1'] = 'id'
    outSheet['B1'] = 'description'
    outSheet['C1'] = 'pictureCount'
    outSheet['D1'] = 'result'
    outSheet['E1'] = 'category'
    outSheet['F1'] = 'severity'
    outSheet['G1'] = 'recurrent'
    # 循环处理表格中每一行数据 并存储
    for row in range(2,rowMax+1):
        id = str(inSheet['A%s' % row].value)
        description = inSheet['D%s' % row].value
        category = inSheet['C%s' % row].value
        severity = inSheet['F%s' % row].value
        recurrent = inSheet['G%s' % row].value
        result = ai(id,description)
        ans = result[1]
        count = result[0]
        outSheet['A%s' % row] = id
        outSheet['B%s' % row] = description
        outSheet['C%s' % row] = count
        outSheet['D%s' % row] = ans
        outSheet['E%s' % row] = category
        outSheet['F%s' % row] = severity
        outSheet['G%s' % row] = recurrent
        print("Process: %i / %i" %(row, rowMax+1))
        # 保存excel
        wb.save('afterFusion.xlsx')
    
    

if __name__ == "__main__":
    main()