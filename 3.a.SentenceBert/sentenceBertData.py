import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from sentence_transformers import SentenceTransformer
from sklearn.cluster import KMeans
import numpy as np
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.manifold import TSNE
import matplotlib.pyplot as plt
from transformers import BertTokenizer, BertModel
import os
import pandas as pd

src_path = '../3.a.SentenceBert/afterPreprocessed.xlsx'
sentences = []
finalGroupsCount = 9 # 只需要在这里改分组数就好了

def main():
    # 资源管理
    inWb = openpyxl.load_workbook(src_path) # 数据源的工作簿
    inSheet = inWb.active

    # 输入处理
    rowMax = inSheet.max_row #数据源Excel最大行数
    colMax = inSheet.max_column #数据源Excel最大列数
    for row in range(1200,rowMax+1):
        descrpition = inSheet['D%s' % row].value
        sentences.append(descrpition)
    print("Input finish")
    # 讲句子转化为向量
    tokenizer = BertTokenizer.from_pretrained('bert-base-chinese')
    model = BertModel.from_pretrained('bert-base-chinese')
    inputs = tokenizer(sentences, padding=True, truncation=True, return_tensors='pt',max_length=128)
    outputs = model(**inputs)
    # 使用[CLS]标记的输出作为句子的嵌入向量
    embeddings = outputs.last_hidden_state
    # 这是一个shape = （1400，128，768）的大向量
    # 将嵌入向量转换为numpy数组
    embeddings = embeddings.detach().numpy()
    # 将列表存储到npy文件
    np.save("embeddings.npy",embeddings)

if __name__ == "__main__":
    main()