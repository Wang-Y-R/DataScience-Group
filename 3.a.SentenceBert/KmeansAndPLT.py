import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from sentence_transformers import SentenceTransformer
from sklearn.cluster import KMeans
import numpy as np
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.manifold import TSNE
import matplotlib.pyplot as plt
from transformers import BertTokenizer, BertModel
import os
from sklearn.decomposition import PCA

# 设置OMP_NUM_THREADS环境变量
os.environ['OMP_NUM_THREADS'] = '1'
src_path = '../3.a.SentenceBert/afterPreprocessed.xlsx'
save_path = 'sentenceBertResult.xlsx'

finalGroupsCount = 14 # 只需要在这里改分组数就好了

def main():
    # 资源管理
    inWb = openpyxl.load_workbook(src_path) # 数据源的工作簿
    inSheet = inWb.active
    outWb = openpyxl.load_workbook(save_path) # 存储的工作簿
    outSheet = outWb.active
    
    # 设置最终输出的Excel的格式
    outSheet['A1'] = 'finalGroup'
    outSheet['B1'] = 'id'
    outSheet['C1'] = 'description'
    outSheet['D1'] = 'pictureCount'
    outSheet['E1'] = 'fusionResult'
    outSheet['F1'] = 'category'
    outSheet['G1'] = 'severity'
    outSheet['H1'] = 'recurrent'
    outSheet['I1'] = 'preProcessedResult'
    
    # 输入处理
    rowMax = inSheet.max_row #数据源Excel最大行数
    colMax = inSheet.max_column #数据源Excel最大列数
    print("There are %i rows."%(rowMax))
    for row in range(2,rowMax+1):
        descrpition = inSheet['D%s' % row].value
        outSheet['B%s' % row] = inSheet['A%s' % row].value
        outSheet['C%s' % row] = inSheet['B%s' % row].value
        outSheet['D%s' % row] = inSheet['C%s' % row].value
        outSheet['E%s' % row] = inSheet['D%s' % row].value
        outSheet['F%s' % row] = inSheet['E%s' % row].value
        outSheet['G%s' % row] = inSheet['F%s' % row].value
        outSheet['H%s' % row] = inSheet['G%s' % row].value
        outSheet['I%s' % row] = inSheet['H%s' % row].value

    # 从npy文件读入句子向量数据
    embeddings1 = np.load('embeddings1.npy')
    embeddings2 = np.load('embeddings2.npy')
    embeddings3 = np.load('embeddings3.npy')
    embeddings4 = np.load('embeddings4.npy')
    embeddings5 = np.load('embeddings5.npy')
    embeddings = np.concatenate((embeddings1, embeddings2,embeddings3,embeddings4,embeddings5), axis=0)
    
    # 降低维度
    flattened_embeddings = embeddings.reshape(-1, 768)
    print("finish flatten")
    
    # 拟合数据并应用PCA变换
    print("pca start")
    pca = PCA(n_components=0.9)
    pca.fit(flattened_embeddings)
    transformed_embeddings = pca.transform(flattened_embeddings)
    
    # 使用K-means聚类
    print("k-means start")
    num_clusters = finalGroupsCount
    kmeans = KMeans(n_clusters=num_clusters,n_init= 10)
    kmeans.fit(transformed_embeddings)
    clusters = kmeans.labels_
    print("finish k-means")
    




# 选择前两个主成分作为特征
    feature1 = transformed_embeddings[:, 0]
    feature2 = transformed_embeddings[:, 1]

# 绘制 t-SNE 结果的散点图
    plt.figure(figsize=(10, 6))
    scatter = plt.scatter(feature1, feature2, c=clusters, cmap='rainbow')
    plt.colorbar(scatter)
    plt.title('t-SNE Visualization with K-means Clustering')
    path = 'sentenceBertResult' + str(finalGroupsCount) + '.png'
    plt.savefig(path,dpi = 100)
    plt.show()


    # 讲结果存储到Excel
    for i in range(2,rowMax+1):
        outSheet['A%s' % i] = clusters[i]
        
    # 保存
    outWb.save('sentenceBertResult' + str(finalGroupsCount) + '.xlsx')
    
if __name__ == "__main__":
    main()