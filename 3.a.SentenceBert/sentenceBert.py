import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from sentence_transformers import SentenceTransformer
from sklearn.cluster import KMeans
import numpy as np
# 加载Sentence-Bert模型
model = SentenceTransformer('all-MiniLM-L6-v2')

src_path = '../SentenceBert/afterPreprocessed.xlsx'
save_path = 'sentenceBertResult.xlsx'
sentences = []

finalGroupsCount = 25

def main():
    # 资源管理
    inWb = openpyxl.load_workbook(src_path) # 数据源的工作簿
    inSheet = inWb.active
    outWb = openpyxl.load_workbook(save_path) # 存储的工作簿
    outSheet = outWb.active
    
    # 设置最终输出的Excel的格式
    outSheet['A1'] = 'finalGroup'
    outSheet['B1'] = 'id'
    outSheet['C1'] = 'description'
    outSheet['D1'] = 'pictureCount'
    outSheet['E1'] = 'fusionResult'
    outSheet['F1'] = 'category'
    outSheet['G1'] = 'severity'
    outSheet['H1'] = 'recurrent'
    outSheet['I1'] = 'preProcessedResult'
    
    # 输入处理
    rowMax = inSheet.max_row #数据源Excel最大行数
    colMax = inSheet.max_column #数据源Excel最大列数
    print("There are %i rows."%(rowMax))
    for row in range(2,rowMax+1):
        descrpition = inSheet['H%s' % row].value
        outSheet['B%s' % row] = inSheet['A%s' % row].value
        outSheet['C%s' % row] = inSheet['B%s' % row].value
        outSheet['D%s' % row] = inSheet['C%s' % row].value
        outSheet['E%s' % row] = inSheet['D%s' % row].value
        outSheet['F%s' % row] = inSheet['E%s' % row].value
        outSheet['G%s' % row] = inSheet['F%s' % row].value
        outSheet['H%s' % row] = inSheet['G%s' % row].value
        outSheet['I%s' % row] = inSheet['H%s' % row].value
        sentences.append(descrpition)
    
    # 讲句子转化为向量
    sentence_embeddings = model.encode(sentences)
    print(sentence_embeddings)
    # 使用K-means聚类
    num_clusters = finalGroupsCount
    kmeans = KMeans(n_clusters=num_clusters)
    kmeans.fit(sentence_embeddings)
    clusters = kmeans.labels_

    # 打印聚类结果
    # for i, sentence in enumerate(sentences):
    #     print(f"Sentence: '{sentence}' - Cluster: {clusters[i]}")

    # 讲结果存储到Excel
    for i, sentence in enumerate(sentences):
        outSheet['A%s' %(i+2)] = clusters[i]
        
    # 保存
    outWb.save('sentenceBertResult' + str(finalGroupsCount) + '.xlsx')
    
if __name__ == "__main__":
    main()