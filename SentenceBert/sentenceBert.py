import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import os
from sentence_transformers import SentenceTransformer
from sklearn.cluster import KMeans
import numpy as np

# 加载Sentence-Bert模型
model = SentenceTransformer('all-MiniLM-L6-v2')

src_path = '../GraphAndTextFusion/afterFusion.xlsx'
save_path = 'sentenceBertResult.xlsx'

sentences = []

def sentenceInput(inSheet):
        # 读入句子
    rowMax = inSheet.max_row #数据源Excel最大行数
    colMax = inSheet.max_column #数据源Excel最大列数
    print("There are %i rows."%(rowMax))
    for row in range(2,rowMax+1):
        descrpition = inSheet['D%s' % row].value
        sentences.append(descrpition)
    
def main():
    # 资源管理
    inWb = openpyxl.load_workbook(src_path) # 数据源的工作簿
    inSheet = inWb.active
    outWb = openpyxl.load_workbook(save_path) # 存储的工作簿
    outSheet = outWb.active
    
    # 设置最终输出的Excel的格式
    outSheet['A1'] = 'finalGroup'
    outSheet['B1'] = 'id'
    outSheet['C1'] = 'description'
    outSheet['D1'] = 'result'
    outSheet['E1'] = 'category'
    outSheet['F1'] = 'severity'
    outSheet['G1'] = 'recurrent'
    
    # 输入处理的句子
    sentenceInput(inSheet)
    
    # 讲句子转化为向量
    sentence_embeddings = model.encode(sentences)
    
    # 使用K-means聚类
    num_clusters = 100
    kmeans = KMeans(n_clusters=num_clusters)
    kmeans.fit(sentence_embeddings)
    clusters = kmeans.labels_

    # 打印聚类结果
    # for i, sentence in enumerate(sentences):
    #     print(f"Sentence: '{sentence}' - Cluster: {clusters[i]}")

    # 讲结果存储到Excel
    for i, sentence in enumerate(sentences):
        outSheet['A%s' %(i+2)] = clusters[i]
        outSheet['C%s' %(i+2)] = sentence
        
    # 保存
    outWb.save('sentenceBertResult.xlsx')
    
if __name__ == "__main__":
    main()