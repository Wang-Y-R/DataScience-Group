import os
from openpyxl import load_workbook

# 获取当前目录下的所有xlsx文件
xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]

# 遍历xlsx文件列表
for file in xlsx_files:
    # 加载工作簿
    wb = load_workbook(filename=file)

    # 获取文件的基本名，不包含扩展名
    base_name = os.path.splitext(file)[0]

    # 重命名第一个工作表为文件的基本名
    if wb.sheetnames[0] != base_name:
        wb.active = wb[wb.sheetnames[0]]
        wb.active.title = base_name

    # 检查是否有第二个工作表，如果有，则重命名
    if len(wb.sheetnames) > 1:
        # 重命名第二个工作表，添加前缀
        second_sheet_name = wb.sheetnames[1]
        new_name = f"{base_name}_WF"
        wb[second_sheet_name].title = new_name

    # 保存工作簿
    wb.save(filename=file)
    print(f'Updated "{file}"')