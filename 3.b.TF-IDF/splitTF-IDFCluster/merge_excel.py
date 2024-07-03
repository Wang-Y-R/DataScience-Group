import os
from openpyxl import load_workbook, Workbook

# 创建一个新的工作簿对象，用于保存合并后的工作表
merged_wb = Workbook()

# 获取当前目录下的所有xlsx文件
xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]

# 遍历xlsx文件列表
for file in xlsx_files:
    # 加载工作簿
    wb = load_workbook(filename=file)

    # 遍历工作簿中的所有工作表
    for sheet in wb.sheetnames:
        # 从当前工作簿复制工作表到新的工作簿
        sheet_obj = wb[sheet]
        # 复制工作表时，如果新工作簿中已存在同名工作表，将重命名
        new_sheet_name = f"{sheet}"
        while new_sheet_name in merged_wb.sheetnames:
            new_sheet_name += "_Copy"

        # 将工作表添加到新工作簿
        merged_wb.create_sheet(title=new_sheet_name)
        for row in sheet_obj.iter_rows():
            for cell in row:
                merged_wb[new_sheet_name][cell.coordinate].value = cell.value

# 保存合并后的工作簿到clusters.xlsx
merged_wb.save('clusters.xlsx')
print('All sheets have been merged into clusters.xlsx')